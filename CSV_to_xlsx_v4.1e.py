import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(script_dir)

print("==============================================================")
print("                TEMU Produkt-Update Automatisierung           ")
print("==============================================================\n")

print("Achte darauf, dass KEINE der betroffenen Dateien geöffnet")
print("oder schreibgeschützt sind!\n")

print("TUTORIAL:")
print("- Exportiere die CSV über Plenty mit entsprechendem Filter")
print("- Führe die Python-Batch-Datei aus")
print("- Warte auf die Bestätigung des Skripts")
print("- Importiere die fertige Excel-Datei bei Temu")
print("- Kategorie anpassen [Automatisierung aktiv] -> Fertig!\n")

print("==============================================================")
input("Drücke ENTER um fortzufahren ...")
print("==============================================================\n")


try:
    # ==============================================================
    # 1. Grundkonfiguration
    # ==============================================================

    csv_path = "Temu_HasCat_IsTemu.csv"   # Pfad zu deiner CSV
    excel_path = "TEMU.xlsx"             # Pfad zu deiner bestehenden xlsx
    sheet_name = "Template"              # Tabellenblattname
    start_row = 5                        # Ab dieser Zeile wird in Excel geschrieben

    # ==============================================================
    # 2. CSV einlesen (Semikolon-separiert)
    # ==============================================================

    print("[INFO] Lese CSV-Datei ein ...")
    df = pd.read_csv(
        csv_path,
        sep=";",
        encoding="utf-8",
        dtype=str
    )

    print("[INFO] Spalten in der CSV:")
    print(list(df.columns))
    print()

    # ==============================================================
    # Hilfsfunktionen
    # ==============================================================

    def clean_richtext(value):
        """Entfernt HTML/Markdown/Rich-Text-Reste, reduziert Whitespaces."""
        if pd.isna(value):
            return ""
        text = str(value)

        # HTML-Tags entfernen
        text = re.sub(r"<.*?>", "", text)

        # Markdown: Fett/Kursiv/Liste/Heading entfernen
        text = re.sub(r"[*_`#>]", "", text)

        # HTML Entities (&nbsp; &amp; etc.) entfernen
        text = re.sub(r"&[a-zA-Z0-9#]+;", " ", text)

        # Mehrere Leerzeichen reduzieren
        text = re.sub(r"\s+", " ", text)

        return text.strip()

    def limit_length(value, max_length):
        if value is None:
            return ""
        text = str(value)
        return text[:max_length] if len(text) > max_length else text

    def excel_col_range(start_col_letter, end_col_letter):
        start_idx = column_index_from_string(start_col_letter)
        end_idx = column_index_from_string(end_col_letter)
        return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]

    def parse_bullets(raw):
        """
        Konvertiert HTML-Aufzählungen aus CSV (<li>, <br>) in eine Liste einzelner Bullet-Zeilen.
        """
        if pd.isna(raw) or raw is None:
            return []
        text = str(raw)

        # HTML Linebreaks / List-Items zu Zeilen
        text = re.sub(r"(?i)<br\s*/?>", "\n", text)     # <br>, <br/>, <BR>
        text = re.sub(r"(?i)</li\s*>", "\n", text)      # </li> -> neue Zeile
        text = re.sub(r"(?i)<li\s*>", "", text)         # <li> entfernen

        # Restliche Tags entfernen
        text = re.sub(r"<.*?>", "", text)

        lines = [ln.strip() for ln in text.split("\n")]
        return [ln for ln in lines if ln]

    def parse_detail_images(raw):
        """
        Konvertiert Detailbilder-String aus CSV:
        'url;0,url;1,' -> ['url', 'url']
        """
        if pd.isna(raw) or raw is None:
            return []
        text = str(raw).strip()
        if not text:
            return []

        items = []
        for part in text.split(","):
            part = part.strip()
            if not part:
                continue
            url = part.split(";")[0].strip()
            if url:
                items.append(url)
        return items

    # ==============================================================
    # 2a. Standardwerte / Defaults
    # ==============================================================

    # Anzahl -> 1 falls leer/fehlend
    col_anzahl = "Anzahl"
    if col_anzahl in df.columns:
        df[col_anzahl] = df[col_anzahl].fillna("1").replace("", "1")
        print("[INFO] Leere Werte in 'Anzahl' wurden auf 1 gesetzt.\n")
    else:
        df[col_anzahl] = "1"
        print("[INFO] Spalte 'Anzahl' fehlte und wurde mit Standardwert 1 angelegt.\n")

    # pandas N/A workaround
    col_nv_listenpreis = "Nicht verfügbar für Listenpreis"
    if col_nv_listenpreis in df.columns:
        df[col_nv_listenpreis] = df[col_nv_listenpreis].fillna("N/A")
        print("[INFO] Leere Werte in 'Nicht verfügbar für Listenpreis' wurden zu 'N/A' gesetzt.\n")
    else:
        print("[WARNUNG] Spalte 'Nicht verfügbar für Listenpreis' nicht in der CSV gefunden.\n")

    # ==============================================================
    # 3. Übersetzung Ursprungsland & Umrechnen
    # ==============================================================

    country_map = {
        "Taiwan": "TW",
        # "China": "CN",
        # "Deutschland": "DE",
    }

    col_country = "Ursprungsland/-region"
    if col_country in df.columns:
        def translate_country(value):
            if pd.isna(value):
                return value
            text = str(value).strip()
            return country_map.get(text, text)

        df[col_country] = df[col_country].apply(translate_country)
        print("[INFO] Spalte 'Ursprungsland/-region' wurde gemäß Mapping übersetzt.\n")
    else:
        print("[WARNUNG] Spalte 'Ursprungsland/-region' nicht in der CSV gefunden.\n")

    # Umrechnen mm -> cm + Fallback 0
    print("[INFO] Konvertiere Maße von mm in cm (Länge/Breite/Höhe) ...")

    mm_to_cm_columns = ["Länge - mm", "Breite - mm", "Höhe - mm"]

    for col in mm_to_cm_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = (df[col] / 10).fillna(0)  # mm -> cm, fallback 0
            print(f"  -> Spalte '{col}' wurde von mm in cm umgerechnet (Fallback 0).")
        else:
            print(f"  -> Hinweis: Spalte '{col}' ist in der CSV nicht vorhanden.")

    print()

    # ==============================================================
    # 4. Stückzahlen aus "Gesamtartikelanzahl" auslesen und berechnen
    # ==============================================================

    print("[INFO] Berechne Gesamtartikelanzahl aus Artikel-Stückzahlen ...")

    source_col = "Gesamtartikelanzahl"
    target_col = "Gesamtartikelanzahl"
    target_col = "Artikel"

    def sum_stueckzahlen(value):
        """
        Erwartet ein Format wie: 'xxxx:1;xxxx:2;xxxx:1'
        Gibt die Summe der Mengen zurück, z.B. 4
        """
        if pd.isna(value):
            return 0

        try:
            parts = str(value).split(";")
            total = 0
            for part in parts:
                if ":" in part:
                    _, qty = part.split(":")
                    qty = qty.strip()
                    if qty:
                        total += int(qty)
            return 1 if total < 1 else total
        except Exception:
            return 1

    if source_col in df.columns:
        if target_col not in df.columns:
            df[target_col] = 0
        df[target_col] = df[source_col].apply(sum_stueckzahlen)
        print(f"[INFO] Spalte '{target_col}' wurde aus '{source_col}' berechnet.\n")
    else:
        print(f"[WARNUNG] Spalte '{source_col}' nicht gefunden oder leer – Gesamtartikelanzahl wird nicht berechnet.\n")

    # ==============================================================
    # 5. Kategorie-Automatisierung per Keywords (Fallback: unverändert)
    # ==============================================================

    category_col = "Kategorie"

    # Priorität: spezifischer zuerst (oben = gewinnt bei Mehrfachtreffern)
    category_rules = [
        (21842, "Scheinwerferbaugruppen, Teile & Zubehör / Scheinwerferbaugruppen",
         [r"\bscheinwerfer\b", r"\bheadlight\b", r"\bfrontscheinwerfer\b", r"\bleuchtmittel\b"]),
        (21792, "Kühlergrills",
         [r"\bkühlergrill\b", r"\bgrill\b", r"\bkühler\s*grill\b", r"\bradiator\s*grill\b"]),
        (21873, "Spiegel & Teile / Außenspiegel",
         [r"\baußenspiegel\b", r"\bseitenspiegel\b", r"\bspiegelkappen\b", r"\bspiegelkappe\b", r"\bspiegel\b", r"\bmirror\b"]),
        (22226, "Frontspoiler",
         [r"\bfrontspoiler\b", r"\bfrontlippe\b", r"\bsplitter\b", r"\bfront\s*lip\b"]),
        (22229, "Spoiler",
         [r"\bspoiler\b", r"\bhec?kspoiler\b", r"\bheckflügel\b", r"\bwing\b", r"\bheckspoiler\b"]),
        (21775, "Stoßdämpfer",
         [r"\bstoßdämpfer\b", r"\bshock\s*absorber\b", r"\bgewindefahrwerk\b"]),
        (21691, "Dämpfer",
         [r"\bdämpfer\b", r"\bdamper\b"]),
        (21903, "Innenraum / Innenraumschalter",
         [r"\binnenraumschalter\b", r"\bschalter\b", r"\bwindow\s*switch\b", r"\bknopf\b",
          r"\bbedienelement\b", r"\binnenraum\b"]),
        (21680, "Abgase & Emissionen / Auspuffrohre & -Endrohre",
         [r"\bauspuff\b", r"\bflexrohr\b", r"\bendrohr\b", r"\babgas\b", r"\brohrschelle\b", r"\brohrverbinder\b"]),
        (21785, "Karosserie & Zierleisten / Körper / Karosseriesätze",
         [r"\bdomstrebe\b", r"\bzierleiste\b", r"\bkantenschutz\b", r"\bschürze\b",
          r"\bkarosserie\b", r"\bblende\b"]),
        # (21226, "Beleuchtung & Elektrik / Beleuchtung",
         # [r"\bbeleuchtung\b", r"\bblinker\b", r"\brücklicht\b", r"\bbremslicht\b", r"\bnebellicht\b",
          # r"\bkennzeichenleuchte\b", r"\belektrik\b"]),
    ]

    def pick_category_from_text(text):
        if not text:
            return None
        t = str(text)
        for cat_id, _, patterns in category_rules:
            for pat in patterns:
                if re.search(pat, t, flags=re.IGNORECASE):
                    return str(cat_id)
        return None

    # Textbasis für Matching (Beschreibung + Name + Bullets)
    def build_match_text(row):
        parts = []
        for c in ["Produktbeschreibung", "Produktname", "Aufzählungspunkt"]:
            if c in df.columns:
                v = row.get(c, "")
                if pd.isna(v):
                    v = ""
                parts.append(str(v))
        return " ".join(parts)

    if category_col in df.columns:
        assigned = 0
        for idx, row in df.iterrows():
            match_text = build_match_text(row)
            cat = pick_category_from_text(match_text)
            if cat:
                df.at[idx, category_col] = cat
                assigned += 1
        print(f"[INFO] Kategorie-Automatisierung: {assigned} Zeilen per Keyword-Klassifizierung gesetzt.\n")
    else:
        print("[WARNUNG] Spalte 'Kategorie' nicht in der CSV gefunden – Automatisierung übersprungen.\n")

    # ==============================================================
    # 5.1 Produkte ohne Produktidentifikation und ohne Preis ignorieren
    # ==============================================================

    col_pid = "Produktidentifikation"
    col_pid = "Listenpreis - EUR"

    ignored_count = 0
    before_count = len(df)

    if col_pid in df.columns:
         # Leerwerte normalisieren
        pid = df[col_pid].fillna("").astype(str).str.strip()

        # Optional: auch "N/A" und "-" als leer behandeln
        is_missing = (pid == "") | (pid.str.lower().isin(["999.99", "n/a", "na", "-", "null", "none"]))

        ignored_count = int(is_missing.sum())
        df = df.loc[~is_missing].copy()

        print(f"[INFO] Fehlerhafte Produkte ignoriert: {ignored_count} von {before_count}")
        print(f"[INFO] Verbleibende Produkte zur Übertragung: {len(df)}\n")
    else:
        print("[WARNUNG] Spalte 'Produktidentifikation' nicht in der CSV gefunden – Filterung übersprungen.\n")

    # ==============================================================
    # 6. Excel-Vorlage öffnen
    # ==============================================================

    print("[INFO] Öffne Excel-Datei ...")
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt '{sheet_name}' wurde nicht in der Excel-Datei gefunden.")
    ws = wb[sheet_name]
    print(f"[INFO] Tabellenblatt '{sheet_name}' geladen.\n")

    # ==============================================================
    # 7. Spalten-Mapping (CSV -> Excel)
    #    Aufzählungspunkt: T–Y (Zeilen)
    #    URL für Detailbilder: Z–BU (Zeilen)
    # ==============================================================

    column_mapping = {
        "Warennummer": "A",
        "SKU-ID": "B",
        "Status": "C",
        "Details": "D",
        "Kategorie": "E",
        "Kategoriename": "F",
        "Produkttyp": "G",
        "Produktname": "L",
        "Verkäufer Goods": "M",
        "Verkäufer SKU": "N",
        "Aktualisieren oder hinzufügen": "O",
        "Marke": "R",
        "Markenzeichen": "S",
        "Produktbeschreibung": "T",

        # Multi-Cell Felder:
        "Aufzählungspunkt": excel_col_range("U", "Z"),
        "URL für Detailbilder": excel_col_range("AA", "BX"),

        "1081 - Oberflächenbehandlungsprozess": "CH",
        "1279 - Verwendungszweck": "CM",
        "Variationsdesign": "DA",
        "Farbe": "DB",
        "Größe": "DC",
        "Stil": "DD",
        "Material": "DE",
        "Aromen": "DF",
        "Anwendbare Personen": "DG",
        "Kapazität": "DH",
        "Komposition": "DI",
        "Gewicht": "DJ",
        "Artikel": "DK",
        "Menge": "DL",
        "Modell": "DM",
        "Haarlänge": "DN",
        "URL für SKU-Bilder": "DO",
        "Abmessungen": "DY",
        "Anzahl": "DZ",
        "Basispreis - EUR": "EA",
        "Referenzlink": "EB",
        "Listenpreis - EUR": "EC",
        "Nicht verfügbar für Listenpreis": "ED",
        "Gewicht des Pakets - g": "EE",
        "Länge - mm": "EF",
        "Breite - mm": "EG",
        "Höhe - mm": "EH",
        "SKU-Typ": "EI",
        "Einzeln verpackt": "EJ",
        "Gesamtverpackungsanzahl": "EK",
        "Verpackungseinheit": "EL",
        # "Artikeleinheit": "DS",
        "Nettoinhalt": "EO",
        "Gesamtnettoinhalt": "EP",
        "Nettoinhaltseinheit": "EQ",
        "Versandvorlage": "ET",
        "Ursprungsland/-region": "EX",
        "Herkunftsregion": "EY",
        "SKU-Verpackungsinformationen (mit sichtbarem Etikett)": "EZ",
        "Ursprungsetikett & Herstellerinformationen": "GS",
        "Wurden Produkte unter dieser Waren-ID nach dem 13. Dezember 2024 in der EU (oder Nordirland) platziert?": "GY",
        "Produktidentifikation": "GZ",
        "Hersteller": "HA",
        "EU-Verantwortlicher": "HB",
    }

    # ==============================================================
    # 8. Alte Daten ab start_row in den relevanten Spalten leeren
    # ==============================================================

    print("[INFO] Leere alte Daten in der Excel-Datei ab Zeile", start_row, "...")

    max_row = ws.max_row

    # Mapping-Spalten flach machen (auch Listenbereiche)
    used_excel_cols = []
    for v in column_mapping.values():
        if not v:
            continue
        if isinstance(v, list):
            used_excel_cols.extend(v)
        else:
            used_excel_cols.append(v)

    # Deduplizieren
    used_excel_cols = list(dict.fromkeys(used_excel_cols))

    for excel_col in used_excel_cols:
        if isinstance(excel_col, int):
            col_index = excel_col
            for row in range(start_row, max_row + 1):
                ws.cell(row=row, column=col_index, value=None)
        else:
            for row in range(start_row, max_row + 1):
                ws[f"{excel_col}{row}"] = None

    print("[INFO] Alte Daten wurden entfernt.\n")

    # ==============================================================
    # 9. Daten aus der CSV in das Excel-Sheet schreiben
    # ==============================================================

    print("[INFO] Schreibe neue Daten in die Excel-Datei ...")

    for i, (_, data_row) in enumerate(df.iterrows(), start=start_row):
        for csv_col, excel_col in column_mapping.items():
            if not excel_col:
                continue
            if csv_col not in df.columns:
                continue

            raw_value = data_row[csv_col]

            # --- Multi-Spalten-Felder ---
            if isinstance(excel_col, list):
                if csv_col == "Aufzählungspunkt":
                    values = parse_bullets(raw_value)
                elif csv_col == "URL für Detailbilder":
                    values = parse_detail_images(raw_value)
                else:
                    values = []

                for j, col_letter in enumerate(excel_col):
                    v = values[j] if j < len(values) else ""
                    v = clean_richtext(v)  # erst nach dem Zerlegen bereinigen

                    # Temu Limit: max 700 Zeichen je Aufzählungspunkt-Zelle
                    if csv_col == "Aufzählungspunkt":
                        v = limit_length(v, 700)

                    ws[f"{col_letter}{i}"] = v

                continue  # nicht nochmal "normal" schreiben

            # --- Normale 1:1 Felder ---
            value = clean_richtext(raw_value)

            # Temu Limit: URL für SKU-Bilder max 512 Zeichen
            if csv_col == "URL für SKU-Bilder":
                value = limit_length(value, 512)

            # (optional) weitere Limits leicht erweiterbar:
            # if csv_col == "Produktbeschreibung":
            #     value = limit_length(value, 5000)

            if isinstance(excel_col, int):
                ws.cell(row=i, column=excel_col, value=value)
            else:
                ws[f"{excel_col}{i}"] = value

    print("[INFO] Schreiben abgeschlossen.\n")

    # ==============================================================
    # 10. Excel-Datei speichern
    # ==============================================================

    wb.save(excel_path)
    print("==============================================================")
    print("[FERTIG] Die Excel-Datei wurde erfolgreich aktualisiert.")
    print(f"         Datei: {excel_path}")
    print("==============================================================\n")

except Exception as e:
    print("\n--- FEHLER AUFGETRETEN ----------------------------------------")
    print(e)
    print("----------------------------------------------------------------\n")

finally:
    input("Weiter mit ENTER ...")

import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# Sicherstellen, dass das Skript-Verzeichnis als aktuelles Verzeichnis gesetzt ist
script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(script_dir)

print("==============================================================")
print("              TEMU Produkt-Update Automatisierung             ")
print("==============================================================\n")

print("Achte darauf, dass **keine** der betroffenen Dateien geöffnet oder schreibgeschützt sind!\n")

print("TUTORIAL:")
print("- Exportiere die CSV über Plenty mit entsprechendem Filter")
print("- Führe die Python-Batch-Datei aus")
print("- Warte auf die Bestätigung des Skripts")
print("- Importiere die fertige Excel-Datei bei Temu")
print("- Kategorie prüfen/anpassen (Automatisierung aktiv) – fertig!\n")

print("==============================================================")
input("Drücke ENTER um fortzufahren ...")
print("==============================================================\n")

try:
    # 1. Grundkonfiguration
    csv_path = "Temu_HasCat_IsTemu.csv"   # Pfad zur CSV-Datei
    excel_path = "TEMU.xlsx"             # Pfad zur bestehenden Excel-Vorlage
    sheet_name = "Template"              # Tabellenblattname in der Vorlage
    start_row = 5                        # Ab dieser Zeile werden Daten in Excel geschrieben

    # 2. CSV einlesen (Semikolon-separiert)
    print("Lese CSV-Datei ein ...")
    df = pd.read_csv(csv_path, sep=";", encoding="utf-8", dtype=str)

    # Spaltenübersicht ausgeben
    columns = list(df.columns)
    print("Spalten in der CSV:", ", ".join(columns), "\n")

    # Hilfsfunktionen
    def clean_richtext(value):
        """Entfernt HTML-/Markdown-Tags und reduziert überflüssige Leerzeichen."""
        if pd.isna(value):
            return ""
        text = str(value)
        # HTML-Tags entfernen
        text = re.sub(r"<.*?>", "", text)
        # Markdown-Syntax (fett/kursiv/Listen/Heading) entfernen
        text = re.sub(r"[*`#>]", "", text)
        # HTML-Entities (&nbsp; &amp; etc.) durch Leerzeichen ersetzen
        text = re.sub(r"&[a-zA-Z0-9#]+;", " ", text)
        # Mehrfache Leerzeichen reduzieren
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def limit_length(value, max_length):
        """Begrenzt die Textlänge auf max_length Zeichen."""
        if value is None:
            return ""
        text = str(value)
        return text[:max_length] if len(text) > max_length else text

    def excel_col_range(start_col_letter, end_col_letter):
        """Gibt eine Liste von Spaltenbuchstaben von start_col_letter bis end_col_letter zurück (inklusive)."""
        start_idx = column_index_from_string(start_col_letter)
        end_idx = column_index_from_string(end_col_letter)
        return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]

    def parse_bullets(raw):
        """
        Konvertiert HTML-Aufzählungen (<li>, <br>) in eine Liste einzelner Bulletpoint-Zeilen.
        """
        if pd.isna(raw) or raw is None:
            return []
        text = str(raw)
        # HTML-Zeilenumbrüche und Listenelemente in Zeilenumbrüche umwandeln
        text = re.sub(r"(?i)<br\s*/?>", "\n", text)     # <br> zu newline
        text = re.sub(r"(?i)</li\s*>", "\n", text)      # </li> zu newline
        text = re.sub(r"(?i)<li\s*>", "", text)         # führendes <li> entfernen
        # Restliche HTML-Tags entfernen
        text = re.sub(r"<.*?>", "", text)
        # In Zeilen splitten und trimmen
        lines = [ln.strip() for ln in text.split("\n")]
        return [ln for ln in lines if ln]  # leere Zeilen entfernen

    def parse_detail_images(raw):
        """
        Zerlegt den Detailbild-String aus der CSV (Format: 'url;0,url;1,...') in eine Liste von URLs.
        """
        if pd.isna(raw) or raw is None:
            return []
        text = str(raw).strip()
        if not text:
            return []
        items = []
        for part in text.split(","):
            part = part.strip()
            if not part:
                continue
            url = part.split(";", 1)[0].strip()
            if url:
                items.append(url)
        return items

    # 2a. Standardwerte für bestimmte Spalten setzen
    # Anzahl: falls leer, auf "1" setzen
    if "Anzahl" in df.columns:
        df["Anzahl"] = df["Anzahl"].fillna("1").replace("", "1")
        print("Leere Werte in 'Anzahl' wurden auf 1 gesetzt.\n")
    else:
        df["Anzahl"] = "1"
        print("Spalte 'Anzahl' fehlte. Sie wurde mit dem Standardwert 1 hinzugefügt.\n")

    # "Nicht verfügbar für Listenpreis": falls leer, auf "N/A" setzen
    if "Nicht verfügbar für Listenpreis" in df.columns:
        df["Nicht verfügbar für Listenpreis"] = df["Nicht verfügbar für Listenpreis"].fillna("N/A")
        print("Leere Werte in 'Nicht verfügbar für Listenpreis' wurden zu 'N/A' geändert.\n")
    else:
        print("Hinweis: Spalte 'Nicht verfügbar für Listenpreis' nicht in der CSV gefunden.\n")

    # 3. Werte konvertieren (z.B. Ursprungsland übersetzen, Maßeinheiten umrechnen)
    # Mapping für Länder (Ursprungsland/-region)
    country_map = {
        "Taiwan": "TW",
        "Deutschland": "Germany",
        "Polen": "Poland",
        "Türkei": "Türkiye"
    }
    if "Ursprungsland/-region" in df.columns:
        df["Ursprungsland/-region"] = df["Ursprungsland/-region"].apply(
            lambda x: country_map.get(str(x).strip(), str(x).strip()) if pd.notna(x) else x
        )
        print("Spalte 'Ursprungsland/-region' nach Vorgaben übersetzt.\n")
    else:
        print("Hinweis: Spalte 'Ursprungsland/-region' nicht in der CSV gefunden.\n")

    # Maße von mm in cm umrechnen (Minimum 1 cm)
    print("Konvertiere Maße von mm in cm (mindestens 1 cm)...")
    mm_to_cm_cols = ["Länge - mm", "Breite - mm", "Höhe - mm"]
    for col in mm_to_cm_cols:
        if col in df.columns:
            values = pd.to_numeric(df[col], errors="coerce") / 10.0
            # Ungültige Werte auf 1 setzen, Minimum 1
            values = values.fillna(1).clip(lower=1)
            # Auf ganze cm runden und als int speichern
            df[col] = values.round(0).astype(int)
            print(f"  -> Spalte '{col}': Umrechnung mm → cm durchgeführt (Min. 1).")
        else:
            print(f"  -> Hinweis: Spalte '{col}' ist in der CSV nicht vorhanden.")
    print()  # Leerzeile

    # 4. Gesamtanzahl Artikel aus Stückzahlen berechnen
    print("Berechne Gesamtartikelanzahl aus einzelnen Stückzahlen...")
    def sum_stueckzahlen(value):
        """
        Erwartet einen String im Format 'xxxx:1;yyyy:2;...'
        Gibt mindestens 1 zurück (Fallback).
        """
        if pd.isna(value) or str(value).strip() == "":
            return 1
        total = 0
        try:
            parts = str(value).split(";")
            for part in parts:
                if ":" not in part:
                    continue
                _, qty = part.split(":", 1)
                qty = qty.strip()
                if qty.isdigit():
                    total += int(qty)
        except Exception:
            return 1  # Fallback bei Fehler
        return total if total >= 1 else 1

    if "Gesamtartikelanzahl" in df.columns:
        df["Artikel"] = df["Gesamtartikelanzahl"].apply(sum_stueckzahlen)
        print("Spalte 'Artikel' wurde aus 'Gesamtartikelanzahl' berechnet.\n")
    else:
        print("Hinweis: Spalte 'Gesamtartikelanzahl' nicht in der CSV gefunden.\n")

    # 5. Kategorie-Automatisierung per Keywords (falls Kategorie-Spalte vorhanden)
    # Reihenfolge = Priorität (spezifisch -> allgemein)

    category_rules = [
        # Licht / Scheinwerfer & Blinker
        (21842, "Scheinwerferbaugruppen, Teile & Zubehör / Scheinwerferbaugruppen",     [r"\b(haupt|front)?scheinwerfer(n)?\b", r"\bhead\s*light(s)?\b", r"\bheadlight(s)?\b", r"\bnebelscheinwerfer(n)?\b", r"\b(tagfahrlicht|tfl)\b", r"\b(blinker|seitenblinker|blinklicht)(n)?\b",
         r"\b(standlicht|rücklicht|ruecklicht)\b", r"\b(leuchtmittel|birne(n)?|bulb(s)?)\b",]),

        # Kühlergrill / Frontgrill (weniger "grill" allein, mehr Kontext)
        (21792, "Kühlergrills", [ r"\b(kühler|kuehler)\s*grill\b", r"\bkühlergrill\b", r"\bfront\s*grill(e)?\b", r"\bfrontgrill\b", r"\b(radiator|radiator)\s*grill(e)?\b",
         r"\bgrillgitter\b", r"\bkühlergitter\b", r"\bgrille\b",]),

        # Außenspiegel
        (21873, "Spiegel & Teile / Außenspiegel", [ r"\b(außen|aussen)spiegel\b", r"\bseitenspiegel\b", r"\bspiegelkappe(n)?\b", r"\bspiegelglas\b", r"\bspiegelblinker\b", r"\b(wing\s*)?mirror(s)?\b",]),

        # Frontspoiler / Frontlippe (sehr spezifisch -> vor "Spoiler")
        (22226, "Frontspoiler",
     [
         r"\bfrontspoiler\b",
         r"\bfront\s*lip(pe)?\b",
         r"\bfrontlippe\b",
         r"\bsplitter\b",
         r"\bspoilerlippe\b",
         r"\b(front|vorder)\s*spoiler\b",
     ]),

        # Spoiler allgemein (Heck/Dach/Wing etc.)
        (22229, "Spoiler",
     [
         r"\bheckspoiler\b",
         r"\bhec?kspoiler\b",
         r"\bheckflügel\b",
         r"\bheckfluegel\b",
         r"\bdachspoiler\b",
         r"\brear\s*spoiler\b",
         r"\bwing\b",
         r"\bspoiler\b",
         r"\bdiffusor\b",
     ]),

        # Fahrwerk / Stoßdämpfer (spezifisch)
        (21775, "Stoßdämpfer",
     [
         r"\b(stoß|stoss)dämpfer(n)?\b",
         r"\bshock\s*absorber(s)?\b",
         r"\bfederbein(e)?\b",
         r"\bgewinde\s*fahrwerk\b",
         r"\bcoilover(s)?\b",
     ]),

        # Dämpfer (allgemeiner Begriff, daher nach Stoßdämpfer)
        (21691, "Dämpfer",
     [
         r"\bdämpfer(n)?\b",
         r"\bdamper(s)?\b",
     ]),

        # Isolierung / Schall- und Wärmedämmung
        (21796, "Isolierung / Schall- und Wärmedämmung",
     [
         r"\b(isolier(ung|material)|isolation)\b",
         r"\b(schall|geräusch|geraeusch)\s*dämm(ung|matte|material)\b",
         r"\b(wärme|waerme)\s*dämm(ung|matte|material)\b",
         r"\bhitze(schutz|schutzmatte)\b",
         r"\bthermo(mat(te|te)|isolierung)\b",
         r"\b(alubutyl|bitumenmatte|dämmvlies)\b",
         r"\bnoise\s*(deadening|insulation)\b",
         r"\bheat\s*(shield|insulation)\b",
         r"\bsound\s*(deadening|proofing)\b",
         r"\bfire\s*wall\s*insulation\b",
         r"\bhood\s*insulation\b",
     ]),

        # Diebstahlschutz / Schlüssel / Zugangssysteme / Fernbedienung / Funk & Fernsteuerung
        (22066, "Diebstahlschutz / Schlüssel / Zugangssysteme / Fernbedienung / Funk & Fernsteuerung",
     [
         # Schlüssel / Key / Fob
         r"\b(schlüssel|schluessel)\b",
         r"\bkey(s)?\b",
         r"\bkey\s*fob(s)?\b",
         r"\bfernbedienung(en)?\b",
         r"\bfunk\s*fernbedienung(en)?\b",
         r"\bremote(\s*control)?\b",
         r"\bremote\s*key\b",
         r"\bzentralverriegelung\b",
         r"\bcentral\s*locking\b",
         r"\btür\s*(öffner|oeffner)\b",
         r"\bdoor\s*(opener|unlock)\b",
     ]),


        # Innenraum / Schalter (NICHT "schalter" alleine)
        (21903, "Innenraum / Innenraumschalter",
     [
         r"\binnenraum(s)?\b",
         r"\binnenraumschalter\b",
         r"\bfensterheber(\s*schalter)?\b",
         r"\bwindow\s*switch\b",
         r"\blichtschalter\b",
         r"\btaster\b",
         r"\bbedienelement(e)?\b",
         r"\bschalterleiste\b",
     ]),

        # Abgase / Auspuffrohre & Endrohre (Flexrohr bleibt hier = Abgas-Flexrohr)
        (21680, "Abgase & Emissionen / Auspuffrohre & -Endrohre",
     [
         r"\bauspuff\b",
         r"\bendrohr(e)?\b",
         r"\babgas\b",
         r"\bflexrohr\b",
         r"\bdownpipe\b",
         r"\brohrschelle(n)?\b",
         r"\brohrverbinder\b",
         r"\bauspuff(rohr)?\b",
     ]),

        # NEU: Schläuche / Flex (gezielt, nicht jedes "Schlauch" matchen)
        (20555, "20555 - Riemen, Schläuche & Riemenscheiben / Schläuche / Flex",
     [
         r"\bflex\s*schlauch\b",
         r"\bflexschlauch\b",
         r"\bflex\s*hose\b",
         r"\bflexible(r|s)?\s*schlauch\b",
         r"\b(kühlwasser|kuehlwasser)\s*schlauch\b",
         r"\b(wasser|luft|ansaug|turbo|unterdruck|kraftstoff)\s*schlauch\b",
         r"\bintercooler\s*schlauch\b",
     ]),

        # Karosserie & Zierleisten / Karosseriesätze (breit, aber auto-spezifisch)
        (21785, "Karosserie & Zierleisten / Karosserie / Karosseriesätze",
     [
         r"\bkarosserie(satz|kit|teile)?\b",
         r"\bbody\s*kit\b",
         r"\bladekantenschutz\b",
         r"\bzierleiste(n)?\b",
         r"\bkantenschutz\b",
         r"\bnummernschildhalter\b",
         r"\bdomstrebe\b",
         r"\bstoßstange\b",
         r"\bstoßfänger\b",
         r"\bstossfaenger\b",
         r"\bschürze\b",
         r"\bschuerze\b",
         r"\bblende(n)?\b",
         r"\bverkleidung\b",
         r"\bvnummernschild\b",
     ]),
]

    def pick_category_from_text(text):
        if not text:
            return None
        for cat_id, _, patterns in category_rules:
            for pat in patterns:
                if re.search(pat, text, flags=re.IGNORECASE):
                    return str(cat_id)
        return None

    if "Kategorie" in df.columns:
        assigned_count = 0
        for idx, row in df.iterrows():
            # Kategorie anhand relevanter Texte zuordnen (Name, Beschreibung, Kategoriename)
            text = ""
            for col in ["Produktbeschreibung", "Produktname", "Kategoriename"]:
                if col in df.columns and pd.notna(row[col]):
                    text += " " + str(row[col])
            cat_id = pick_category_from_text(text)
            if cat_id:
                df.at[idx, "Kategorie"] = cat_id
                assigned_count += 1
        print(f"Kategorie-Automatisierung: Für {assigned_count} Produkte wurde die Kategorie-ID gesetzt.\n")
    else:
        print("Hinweis: Spalte 'Kategorie' nicht in der CSV gefunden – Automatisierung übersprungen.\n")

    # 5.1 Filter: Produkte aus bestimmten Kategorien nicht übertragen (z.B. 'Garten' oder 'Haushalt')
    filter_terms = ["garten", "haushalt"]
    initial_count = len(df)
    if "Kategorie" in df.columns or "Kategoriename" in df.columns:
        mask = pd.Series(False, index=df.index)
        if "Kategorie" in df.columns:
            mask |= df["Kategorie"].fillna("").astype(str).str.lower().str.contains("|".join(filter_terms))
        if "Kategoriename" in df.columns:
            mask |= df["Kategoriename"].fillna("").astype(str).str.lower().str.contains("|".join(filter_terms))
        ignored = int(mask.sum())
        df = df.loc[~mask].copy()
        print(f"Kategorie-Filter: {ignored} von {initial_count} Produkten ignoriert (enthielten 'Garten' oder 'Haushalt').")
        print(f"Verbleibende Produkte zur Übertragung: {len(df)}\n")
    else:
        print("Hinweis: Keine Kategorie-Spalte vorhanden – Kategorie-Filter übersprungen.\n")

    # 5.2 Produkte ohne wichtige Angaben (Identifikation, Preis, Bild) ignorieren
    required_cols = ["Produktidentifikation", "Listenpreis - EUR", "URL für SKU-Bilder"]
    initial_count = len(df)
    if any(col in df.columns for col in required_cols):
        missing_mask = pd.Series(False, index=df.index)
        for col in required_cols:
            if col in df.columns:
                vals = df[col].fillna("").astype(str).str.strip().str.lower()
                missing_mask |= (vals == "") | (vals.isin(["999.99", "n/a", "na", "-", "null", "none"]))
        ignored = int(missing_mask.sum())
        df = df.loc[~missing_mask].copy()
        print(f"Ungültige Produkte (fehlende ID/Preis/Bild) ignoriert: {ignored} von {initial_count}.")
        print(f"Verbleibende Produkte zur Übertragung: {len(df)}\n")
    else:
        print("Hinweis: Spalten für Produktidentifikation/Preis/Bild fehlen – Überprüfung übersprungen.\n")

    # 6. Excel-Vorlage öffnen und Ziel-Tabelle auswählen
    print("Öffne Excel-Datei...")
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt '{sheet_name}' wurde in der Excel-Datei nicht gefunden.")
    ws = wb[sheet_name]
    print(f"Tabellenblatt '{sheet_name}' erfolgreich geladen.\n")

    # 7. Spalten-Mapping von CSV zu Excel-Spalten (Buchstaben)
    column_mapping = {
        "Warennummer": "A",
        "SKU-ID": "B",
        "Status": "C",
        "Details": "D",
        "Kategorie": "E",
        "Kategoriename": "F",
        "Produkttyp": "G",
        "Produktname": "L",
        "Verkäufer Goods": "M",
        "Verkäufer SKU": "N",
        "Aktualisieren oder hinzufügen": "O",
        "Marke": "R",
        "Markenzeichen": "S",
        "Produktbeschreibung": "T",
        # Mehrere Aufzählungspunkte (bis zu 6 Bullet-Points U–Z)
        "Aufzählungspunkt": excel_col_range("U", "Z"),
        # Detailbilder-URLs (AA–BX)
        "URL für Detailbilder": excel_col_range("AA", "BX"),
        "1081 - Oberflächenbehandlungsprozess": "CB",
        "1279 - Verwendungszweck": "CU",
        "Variationsdesign": "DA",
        "Farbe": "DB",
        "Material": "DE",
        "Kapazität": "DH",
        "Gewicht": "DJ",
        "Artikel": "DK",
        "Menge": "DL",
        "Modell": "DM",
        "URL für SKU-Bilder": "DO",
        "Anzahl": "DZ",
        "Basispreis - EUR": "EA",
        "Referenzlink": "EB",
        "Listenpreis - EUR": "EC",
        "Nicht verfügbar für Listenpreis": "ED",
        "Gewicht des Pakets - g": "EE",
        "Länge - mm": "EF",
        "Breite - mm": "EG",
        "Höhe - mm": "EH",
        "SKU-Typ": "EI",
        "Einzeln verpackt": "EJ",
        "Gesamtverpackungsanzahl": "EK",
        "Verpackungseinheit": "EL",
        "Nettoinhalt": "EO",
        "Gesamtnettoinhalt": "EP",
        "Nettoinhaltseinheit": "EQ",
        "Versandvorlage": "ET",
        "Ursprungsland/-region": "EX",
        "Herkunftsregion": "EY",
        "SKU-Verpackungsinformationen (mit sichtbarem Etikett)": "EZ",
        "Ursprungsetikett & Herstellerinformationen": "GS",
        "Wurden Produkte unter dieser Waren-ID nach dem 13. Dezember 2024 in der EU (oder Nordirland) platziert?": "GY",
        "Produktidentifikation": "GZ",
        "Hersteller": "HA",
        "EU-Verantwortlicher": "HB"
    }

    # 8. Alte Daten aus dem Excel-Blatt entfernen (ab start_row)
    print(f"Lösche alte Einträge in Excel ab Zeile {start_row}...")
    max_row = ws.max_row
    # Alle relevanten Spalten ermitteln (einzeln und Bereiche)
    used_cols = []
    for excel_cols in column_mapping.values():
        if not excel_cols:
            continue
        if isinstance(excel_cols, list):
            used_cols.extend(excel_cols)
        else:
            used_cols.append(excel_cols)
    used_cols = list(dict.fromkeys(used_cols))  # Duplikate entfernen
    # Zellen in den ermittelten Spalten leeren
    for col in used_cols:
        if isinstance(col, int):
            for row in range(start_row, max_row + 1):
                ws.cell(row=row, column=col, value=None)
        else:
            for row in range(start_row, max_row + 1):
                ws[f"{col}{row}"] = None
    print("Alte Daten wurden entfernt.\n")

    # 9. Daten aus dem DataFrame in die Excel-Vorlage schreiben
    print("Schreibe neue Daten in die Excel-Datei...")
    for excel_row, (_, data_row) in enumerate(df.iterrows(), start=start_row):
        for csv_col, excel_cols in column_mapping.items():
            if csv_col not in df.columns or not excel_cols:
                continue
            raw_value = data_row[csv_col]
            # Falls Excel-Zielspalte eine Liste (mehrere Spalten) ist
            if isinstance(excel_cols, list):
                if csv_col == "Aufzählungspunkt":
                    values = parse_bullets(raw_value)
                elif csv_col == "URL für Detailbilder":
                    values = parse_detail_images(raw_value)
                else:
                    values = []
                for j, col_letter in enumerate(excel_cols):
                    cell_value = values[j] if j < len(values) else ""
                    cell_value = clean_richtext(cell_value)
                    # Länge von Bulletpoints auf 700 Zeichen begrenzen
                    if csv_col == "Aufzählungspunkt":
                        cell_value = limit_length(cell_value, 700)
                    ws[f"{col_letter}{excel_row}"] = cell_value
            else:
                # Einzelne Spalte
                cell_value = clean_richtext(raw_value)
                # Länge der SKU-Bild URL auf 512 Zeichen begrenzen
                if csv_col == "URL für SKU-Bilder":
                    cell_value = limit_length(cell_value, 512)
                if isinstance(excel_cols, int):
                    ws.cell(row=excel_row, column=excel_cols, value=cell_value)
                else:
                    ws[f"{excel_cols}{excel_row}"] = cell_value
    print("Schreiben der neuen Daten abgeschlossen.\n")

    # 10. Excel-Datei speichern
    wb.save(excel_path)
    print("==============================================================")
    print("FERTIG - Die Excel-Datei wurde erfolgreich aktualisiert.")
    print(f"         Datei: {excel_path}")
    print("==============================================================\n")

except Exception as e:
    # Fehlermeldung ausgeben
    print("\n--- FEHLER ---")
    print(str(e))
    print("--------------\n")

finally:
    input("Weiter mit ENTER ...")

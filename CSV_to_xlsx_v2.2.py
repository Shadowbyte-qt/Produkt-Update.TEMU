import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(script_dir)

print("==============================================================")
print("                TEMU Produkt-Update Automatisierung           ")
print("==============================================================\n")

print("Achte darauf, dass KEINE der betroffenen Dateien geöffnet")
print("oder schreibgeschützt sind!\n")

print("TUTORIAL:")
print("- Exportiere die CSV über Plenty mit entsprechendem Filter")
print("- Führe die Python-Batch-Datei aus")
print("- Warte auf die Bestätigung des Skripts")
print("- Importiere die fertige Excel-Datei bei Temu")
print("- Kategorie anpassen [Automatisierung geplant] -> Fertig!\n")

print("==============================================================")
input("Drücke ENTER um fortzufahren ...")
print("==============================================================\n")


try:
    # ==============================================================
    # 1. Grundkonfiguration
    # ==============================================================

    csv_path = "Temu_HasCat_IsTemu.csv"          # Pfad zu deiner CSV
    excel_path = "TEMU.xlsx"                     # Pfad zu deiner bestehenden xlsx
    sheet_name = "Template"                      # Tabellenblattname
    start_row = 5                                # Ab dieser Zeile wird in Excel geschrieben

    # ==============================================================
    # 2. CSV einlesen (Semikolon-separiert)
    # ==============================================================

    print("[INFO] Lese CSV-Datei ein ...")
    df = pd.read_csv(
        csv_path,
        sep=";",          # Semikolon als Trennzeichen
        encoding="utf-8", # ggf. anpassen, z.B. 'latin-1'
        dtype=str         # <<< NEU: alles zunächst als Text einlesen
    )

    print("[INFO] Spalten in der CSV:")
    print(list(df.columns))
    print()

    # ==============================================================
    # pandas N/A workaround
    # ==============================================================

    col_nv_listenpreis = "Nicht verfügbar für Listenpreis"

    if col_nv_listenpreis in df.columns:
        # Alle NaN/fehlenden Werte in dieser Spalte durch "N/A" ersetzen
        df[col_nv_listenpreis] = df[col_nv_listenpreis].fillna("N/A")
        print("[INFO] Leere Werte in 'Nicht verfügbar für Listenpreis' wurden zu 'N/A' gesetzt.\n")
    else:
        print("[WARNUNG] Spalte 'Nicht verfügbar für Listenpreis' nicht in der CSV gefunden.\n")

    # ==============================================================
    # 3. Übersetzung Ursprungsland & Umrechnen
    # ==============================================================

    country_map = {
        "Taiwan": "TW",
        # "China": "CN",
        # "Deutschland": "DE",
    }

    col_country = "Ursprungsland/-region"

    if col_country in df.columns:
        def translate_country(value):
            if pd.isna(value):
                return value
            text = str(value).strip()
            return country_map.get(text, text)  # wenn nicht im Dict -> Originalwert

        df[col_country] = df[col_country].apply(translate_country)
        print("[INFO] Spalte 'Ursprungsland/-region' wurde gemäß Mapping übersetzt.\n")
    else:
        print("[WARNUNG] Spalte 'Ursprungsland/-region' nicht in der CSV gefunden.\n")

    # Umrechnen

    print("[INFO] Konvertiere Maße von mm in cm (Länge/Breite/Höhe) ...")

    mm_to_cm_columns = [
        "Länge - mm",
        "Breite - mm",
        "Höhe - mm"
    ]

    for col in mm_to_cm_columns:
        if col in df.columns:
            # dtype=str -> erst in Zahl umwandeln
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col] / 10  # mm -> cm
            print(f"  -> Spalte '{col}' wurde von mm in cm umgerechnet.")
        else:
            print(f"  -> Hinweis: Spalte '{col}' ist in der CSV nicht vorhanden.")

    print()

    # ==============================================================
    # 4. Stückzahlen aus "Artikel" auslesen und in Gesamtartikelanzahl schreiben
    # ==============================================================

    print("[INFO] Berechne Gesamtartikelanzahl aus Artikel-Stückzahlen ...")

    source_col = "Gesamtartikelanzahl"
    target_col = "Gesamtartikelanzahl"

    def sum_stueckzahlen(value):
        """
        Erwartet ein Format wie: 'xxxx:1;xxxx:2;xxxx:1'
        Gibt die Summe der Mengen zurück, z.B. 4
        """
        if pd.isna(value):
            return 0

        try:
            parts = str(value).split(";")
            total = 0

            for part in parts:
                if ":" in part:
                    _, qty = part.split(":")
                    qty = qty.strip()
                    if qty:
                        total += int(qty)

            # Mindeststückzahl = 1
            if total < 1:
                return 1

            return total
        except Exception:
            # Im Fehlerfall lieber 1 statt 0
            return 1

    if source_col in df.columns:
        if target_col not in df.columns:
            df[target_col] = 0
        df[target_col] = df[source_col].apply(sum_stueckzahlen)
        print(f"[INFO] Spalte '{target_col}' wurde aus '{source_col}' berechnet.\n")
    else:
        print(f"[WARNUNG] Spalte '{source_col}' nicht gefunden oder leer – Gesamtartikelanzahl wird nicht berechnet.\n")

    # ==============================================================
    # 5. URL-Spalten für Temu bereinigen (NEU)
    # ==============================================================

    print("[INFO] Bereinige Bild-URL-Spalten für Temu ...")

    def clean_image_urls(value):
        """
        Beispieleingang:
        'https://...1111001.jpg;0,https://...1111001-1.jpg;1,'

        -> 'https://...1111001.jpg|https://...1111001-1.jpg'
        """
        if pd.isna(value):
            return ""

        text = str(value).strip()

        # An Komma oder Semikolon splitten
        parts = re.split(r"[;,]+", text)

        # Nur Einträge, die wie URLs aussehen, behalten
        urls = [p.strip() for p in parts if p.strip().startswith("http")]

        # Mit | verbinden (Temu-konform)
        return "|".join(urls)

    url_columns = [
        "URL für Detailbilder",
        "URL für SKU-Bilder"
    ]

    for col in url_columns:
        if col in df.columns:
            df[col] = df[col].apply(clean_image_urls)
            print(f"  -> Spalte '{col}' wurde bereinigt.")
        else:
            print(f"  -> Hinweis: Spalte '{col}' ist in der CSV nicht vorhanden.")

    print()

    # ==============================================================
    # 6. Excel-Vorlage öffnen
    # ==============================================================

    print("[INFO] Öffne Excel-Datei ...")
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt '{sheet_name}' wurde nicht in der Excel-Datei gefunden.")
    ws = wb[sheet_name]
    print(f"[INFO] Tabellenblatt '{sheet_name}' geladen.\n")

    # ==============================================================
    # 7. Spalten-Mapping (CSV -> Excel)
    #    Werte anpassen! Nur Spalten mit zugewiesenem Buchstaben werden geschrieben.
    # ==============================================================

    column_mapping = {
        "Warennummer": "A",
        "SKU-ID": "B",
        "Status": "C",
        "Details": "D",
        "Kategorie": "E",
        "Produkttyp": "F",
        # "Verarbeitungsverfahren zur Anpassung": "G",
        # "Primärverfahren": "H",
        # "Sekundärverfahren": "I",
        # "Sekundärverfahren.1": "J",  # Zweiter Spaltenname
        "Produktname": "K",
        "Verkäufer Goods": "L",
        "Verkäufer SKU": "M",
        "Aktualisieren oder hinzufügen": "N",
        "Marke": "Q",
        "Markenzeichen": "R",
        "Produktbeschreibung": "S",
        "Aufzählungspunkt": "T",
        "URL für Detailbilder": "Z",
        "Variationsdesign": "CZ",
        "Farbe": "DA",
        "Größe": "DB",
        "Stil": "DC",
        "Material": "DD",
        "Aromen": "DE",
        "Anwendbare Personen": "DF",
        "Kapazität": "DG",
        "Komposition": "DH",
        "Gewicht": "DI",
        "Artikel": "DJ",
        "Menge": "DK",
        "Modell": "DL",
        "Haarlänge": "DM",
        "URL für SKU-Bilder": "DN",
        "Abmessungen": "DX",
        "Anzahl": "DY",
        "Basispreis - EUR": "DZ",
        "Referenzlink": "EA",
        "Listenpreis - EUR": "EB",
        "Nicht verfügbar für Listenpreis": "EC",
        "Gewicht des Pakets - g": "ED",
        "Länge - mm": "EE",
        "Breite - mm": "EF",
        "Höhe - mm": "EG",
        "SKU-Typ": "EH",
        "Einzeln verpackt": "EI",
        "Gesamtverpackungsanzahl": "EJ",
        "Verpackungseinheit": "EK",
        # "Gesamtartikelanzahl": "EL",
        # "Artikeleinheit": "DS",
        "Nettoinhalt": "EN",
        "Gesamtnettoinhalt": "EO",
        "Nettoinhaltseinheit": "EP",
        "Versandvorlage": "ES",
        "Ursprungsland/-region": "EW",
        "Herkunftsregion": "EX",
        "SKU-Verpackungsinformationen (mit sichtbarem Etikett)": "EY",
        "Ursprungsetikett & Herstellerinformationen": "GR",
        "Wurden Produkte unter dieser Waren-ID nach dem 13. Dezember 2024 in der EU (oder Nordirland) platziert?": "GX",
        "Produktidentifikation": "GY",
        "Hersteller": "GZ",
        "EU-Verantwortlicher": "HA",
    }

    # ==============================================================
    # 8. Alte Daten ab start_row in den relevanten Spalten leeren
    # ==============================================================

    print("[INFO] Leere alte Daten in der Excel-Datei ab Zeile", start_row, "...")

    max_row = ws.max_row
    used_excel_cols = [v for v in column_mapping.values() if v]

    for excel_col in used_excel_cols:
        if isinstance(excel_col, int):
            col_index = excel_col
            for row in range(start_row, max_row + 1):
                ws.cell(row=row, column=col_index, value=None)
        else:
            for row in range(start_row, max_row + 1):
                ws[f"{excel_col}{row}"] = None

    print("[INFO] Alte Daten wurden entfernt.\n")

    # ==============================================================
    # 9. Daten aus der CSV in das Excel-Sheet schreiben
    # ==============================================================

    print("[INFO] Schreibe neue Daten in die Excel-Datei ...")

    for i, (_, data_row) in enumerate(df.iterrows(), start=start_row):
        for csv_col, excel_col in column_mapping.items():
            # Überspringen, wenn keine Excel-Spalte zugeordnet ist
            if not excel_col:
                continue

            if csv_col not in df.columns:
                # CSV-Spalte existiert nicht -> überspringen
                continue

            value = data_row[csv_col]

            if isinstance(excel_col, int):  # Spaltenindex
                ws.cell(row=i, column=excel_col, value=value)
            else:  # Spaltenbuchstabe
                ws[f"{excel_col}{i}"] = value

    print("[INFO] Schreiben abgeschlossen.\n")

    # ==============================================================
    # 10. Excel-Datei speichern
    # ==============================================================

    wb.save(excel_path)
    print("==============================================================")
    print("[FERTIG] Die Excel-Datei wurde erfolgreich aktualisiert.")
    print(f"         Datei: {excel_path}")
    print("==============================================================\n")

except Exception as e:
    print("\n--- FEHLER AUFGETRETEN ----------------------------------------")
    print(e)
    print("----------------------------------------------------------------\n")

finally:
    input("Weiter mit ENTER ...")
